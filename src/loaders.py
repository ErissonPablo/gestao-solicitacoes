# -*- coding: utf-8 -*-
"""Leitura dos 3 arquivos da operacao de compras.

- rmatr029  : Solicitacoes de Compra (SCs). Formato SpreadsheetML 2003 (XML).
- rmatr052  : Pedidos de Compra (PCs).      Formato SpreadsheetML 2003 (XML).
- DISTRIBUICAO.xlsx, aba 'base' : ledger de distribuicao para compradores.

Os .xls do Protheus NAO sao binarios: sao XML do Excel. Por isso o parser
proprio em vez de pandas.read_excel.
"""
from __future__ import annotations

import xml.etree.ElementTree as ET
from datetime import datetime

import pandas as pd

SS = "{urn:schemas-microsoft-com:office:spreadsheet}"
NS = {"ss": "urn:schemas-microsoft-com:office:spreadsheet"}


# --------------------------------------------------------------------------- #
# SpreadsheetML (rmatr029 / rmatr052)
# --------------------------------------------------------------------------- #
def _row_cells(row) -> list:
    """Le uma <Row> respeitando ss:Index (celulas puladas viram vazio)."""
    cells, col = [], 0
    for cell in row.findall("ss:Cell", NS):
        idx = cell.get(SS + "Index")
        if idx is not None:
            col = int(idx) - 1
        data = cell.find("ss:Data", NS)
        val = data.text if data is not None else ""
        while len(cells) < col:
            cells.append("")
        cells.append(val if val is not None else "")
        col += 1
    return cells


def _dedupe(names: list[str]) -> list[str]:
    seen, out = {}, []
    for n in names:
        n = (n or "").strip()
        if n in seen:
            seen[n] += 1
            out.append(f"{n}.{seen[n]}")
        else:
            seen[n] = 0
            out.append(n)
    return out


def _read_spreadsheetml(path: str, sheet_contains: str) -> pd.DataFrame:
    tree = ET.parse(path)
    root = tree.getroot()
    ws = None
    for w in root.findall("ss:Worksheet", NS):
        if sheet_contains.lower() in (w.get(SS + "Name") or "").lower():
            ws = w
            break
    if ws is None:
        raise ValueError(f"Aba contendo '{sheet_contains}' nao encontrada em {path}")
    rows = ws.find("ss:Table", NS).findall("ss:Row", NS)
    # rows[0] = titulo do relatorio; rows[1] = cabecalho; resto = dados
    header = _dedupe(_row_cells(rows[1]))
    records = []
    width = len(header)
    for r in rows[2:]:
        vals = _row_cells(r)
        if len(vals) < width:
            vals += [""] * (width - len(vals))
        records.append(vals[:width])
    return pd.DataFrame(records, columns=header)


# --------------------------------------------------------------------------- #
# Datas
# --------------------------------------------------------------------------- #
def parse_data(x):
    """Aceita 'dd/mm/yyyy', ISO '2026-01-15T00:00:00', datetime, ou vazio."""
    if x is None:
        return pd.NaT
    if isinstance(x, datetime):
        return pd.Timestamp(x)
    s = str(x).strip()
    if not s or s.replace("/", "").strip() == "":
        return pd.NaT
    for fmt in ("%d/%m/%Y", "%Y-%m-%dT%H:%M:%S", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
        try:
            return pd.Timestamp(datetime.strptime(s, fmt))
        except ValueError:
            continue
    return pd.to_datetime(s, dayfirst=True, errors="coerce")


def parse_num(x) -> float:
    if x is None:
        return 0.0
    s = str(x).strip().replace(".", "").replace(",", ".") if "," in str(x) else str(x).strip()
    try:
        return float(s)
    except ValueError:
        return 0.0


# --------------------------------------------------------------------------- #
# Loaders publicos
# --------------------------------------------------------------------------- #
def load_scs(path: str) -> pd.DataFrame:
    """Solicitacoes de Compra (rmatr029)."""
    df = _read_spreadsheetml(path, "SOLICITA")
    for c in ("DT.EMISSAO", "DT.LIBERACAO", "DT.NECESSIDADE", "DT.EMIS.PC"):
        if c in df.columns:
            df[c] = df[c].map(parse_data)
    for c in ("QTD.SOLICITADA", "PRC.UNIT.", "TOTAL ITEM"):
        if c in df.columns:
            df[c] = df[c].map(parse_num)
    return df


def load_pcs(path: str) -> pd.DataFrame:
    """Pedidos de Compra (rmatr052)."""
    df = _read_spreadsheetml(path, "Pedido")
    for c in ("EMISSAO", "DATA SOLICIT"):
        if c in df.columns:
            df[c] = df[c].map(parse_data)
    for c in ("QUANTIDADE", "QUANT ENTREG", "PRECO.", "VLR.TOTAL"):
        if c in df.columns:
            df[c] = df[c].map(parse_num)
    return df


def load_distribuicao(path: str, sheet: str = "base") -> pd.DataFrame:
    """Ledger de distribuicao (aba 'base' do xlsx de distribuicao)."""
    import openpyxl

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet]
    it = ws.iter_rows(values_only=True)
    header = [str(h).strip() if h is not None else "" for h in next(it)]
    ncols = 9  # a base so usa as 9 primeiras colunas
    header = header[:ncols]
    rows = []
    for row in it:
        if all(v is None for v in row[:ncols]):
            continue
        rows.append(list(row[:ncols]))
    wb.close()
    df = pd.DataFrame(rows, columns=header)
    if "DATA DISTRIBUIÇÃO" in df.columns:
        df["DATA DISTRIBUIÇÃO"] = df["DATA DISTRIBUIÇÃO"].map(parse_data)
    return df


def load_feriados(path: str, sheet: str = "FERIADOS") -> list:
    """Lista de datas de feriado (para dias uteis)."""
    import openpyxl

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet]
    feriados = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row and isinstance(row[0], datetime):
            feriados.append(pd.Timestamp(row[0]).normalize())
    wb.close()
    return feriados

# -*- coding: utf-8 -*-
"""Resolve de onde vem cada arquivo e QUAL e qual - por CONTEUDO, nao pelo nome.

Os relatorios do Protheus saem com nomes diferentes a cada extracao
('rmatr052 (3).xls', 'rmatr052 12-06.xls'...). Em vez de depender do nome,
abrimos o arquivo e identificamos pela assinatura interna:

    sc   -> Solicitacoes (rmatr029): tem 'QTD.SOLICITADA' / 'NUM.SC'
    pc   -> Pedidos (rmatr052):      tem 'PEDIDO COMPRA' / 'QUANT ENTREG'
    dist -> Distribuicao (.xlsx):    tem aba 'base' com 'DATA DISTRIBUICAO'

Quando ha varios arquivos do mesmo tipo, usamos sempre a EXTRACAO MAIS
RECENTE (data interna do relatorio), garantindo que nada e contado em dobro.
"""
from __future__ import annotations

import glob
import io
import os
import re
import unicodedata
from datetime import datetime

HEAD = 262144  # 256 KB - o cabecalho e os parametros ficam no inicio do arquivo


def _sem_acento_upper(s: str) -> str:
    s = "".join(c for c in unicodedata.normalize("NFD", s)
                if unicodedata.category(c) != "Mn")
    return s.upper()


def _head_text(b: bytes) -> str:
    return _sem_acento_upper(b[:HEAD].decode("utf-8", "ignore"))


def classificar(nome: str, b: bytes) -> str | None:
    """Identifica o tipo do arquivo pelo conteudo. Devolve 'sc'|'pc'|'dist'|None."""
    if not b:
        return None
    # xlsx/zip -> so e distribuicao se tiver a aba 'base' com a cara certa
    if b[:2] == b"PK":
        try:
            import openpyxl

            wb = openpyxl.load_workbook(io.BytesIO(b), read_only=True)
            base = next((s for s in wb.sheetnames if s.strip().lower() == "base"), None)
            cabecalho_ok = False
            if base is not None:
                primeira = next(wb[base].iter_rows(values_only=True), ())
                hdr = _sem_acento_upper(" ".join(str(c) for c in primeira if c))
                cabecalho_ok = "RESPONSAVEL" in hdr and "DISTRIBUI" in hdr
            wb.close()
            return "dist" if cabecalho_ok else None
        except Exception:  # noqa: BLE001
            return None
    # SpreadsheetML (XML) -> SC ou PC
    txt = _head_text(b)
    if "<?XML" in txt or "WORKBOOK" in txt:
        if "QUANT ENTREG" in txt or "PEDIDO COMPRA" in txt:
            return "pc"
        if "QTD.SOLICITADA" in txt or ("NUM.SC" in txt and "SOLICITA" in txt):
            return "sc"
    return None


def data_referencia(tipo: str, b: bytes, mtime: float | None) -> float:
    """Timestamp (epoch) para ordenar 'mais recente'. Usa a data interna do
    relatorio do Protheus (Emissao/Dt.Ref) quando existir; senao, o mtime."""
    if tipo in ("sc", "pc"):
        txt = _head_text(b)
        m = re.search(r"(?:EMISSAO|DT\.REF)[:\s]*([0-3]?\d/[0-1]?\d/\d{4})", txt)
        if m:
            try:
                return datetime.strptime(m.group(1), "%d/%m/%Y").timestamp()
            except ValueError:
                pass
    return mtime if mtime is not None else 0.0


def _bytes_para_classificar(c: dict) -> bytes:
    """Bytes suficientes para classificar: completos se ja em memoria, senao o head."""
    if c.get("bytes") is not None:
        return c["bytes"]
    if c.get("head") is not None:
        return c["head"]
    return b""


def _bytes_completos(c: dict) -> bytes:
    if c.get("bytes") is not None:
        return c["bytes"]
    return ler_bytes(c["path"])


def resolver(candidatos: list[dict]) -> dict:
    """candidatos: lista de dicts com 'nome', 'mtime' e ('bytes' ou 'path'/'head').

    Devolve {'sc': bytes|None, 'pc': bytes|None, 'dist': bytes|None,
             'detalhes': {tipo: {'nome', 'ref', 'descartados':[...]}}}.
    Para cada tipo, escolhe a extracao mais recente; as demais sao ignoradas
    (assim nada e contado em dobro).
    """
    baldes: dict[str, list] = {"sc": [], "pc": [], "dist": []}
    for c in candidatos:
        amostra = _bytes_para_classificar(c)
        tipo = classificar(c.get("nome", ""), amostra)
        if tipo:
            ref = data_referencia(tipo, amostra, c.get("mtime"))
            baldes[tipo].append((ref, c))
    saida = {"sc": None, "pc": None, "dist": None, "detalhes": {}}
    for tipo, itens in baldes.items():
        if not itens:
            continue
        itens.sort(key=lambda x: x[0], reverse=True)
        ref, escolhido = itens[0]
        saida[tipo] = _bytes_completos(escolhido)
        saida["detalhes"][tipo] = {
            "nome": escolhido.get("nome", "?"),
            "ref": datetime.fromtimestamp(ref).strftime("%d/%m/%Y") if ref else "?",
            "descartados": [c.get("nome", "?") for _, c in itens[1:]],
        }
    return saida


def candidatos_da_pasta(pasta: str) -> list[dict]:
    """Varre a pasta. Para .xls (XML) le so o cabecalho (classifica leve);
    .xlsx sao lidos inteiros pois o openpyxl precisa do arquivo todo."""
    arqs = []
    for ext in ("*.xls", "*.xlsx", "*.xml"):
        arqs += glob.glob(os.path.join(pasta, ext))
    out = []
    for p in sorted(set(arqs)):
        nome = os.path.basename(p)
        if nome.startswith("~$"):  # temporario do Office
            continue
        try:
            mtime = os.path.getmtime(p)
            with open(p, "rb") as fh:
                ini = fh.read(2)
                fh.seek(0)
                if ini == b"PK":  # xlsx -> precisa do arquivo inteiro
                    out.append({"nome": nome, "bytes": fh.read(), "mtime": mtime})
                else:  # .xls/.xml -> classifica pelo head, le o resto so se escolhido
                    out.append({"nome": nome, "head": fh.read(HEAD),
                                "path": p, "mtime": mtime})
        except (PermissionError, OSError):
            continue
    return out


def ler_bytes(caminho: str) -> bytes:
    with open(caminho, "rb") as fh:
        return fh.read()
